import streamlit as st
import pandas as pd
import requests
import time
import urllib.parse
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
import random
from datetime import datetime

# ============================================================
# 1. CORE FUNCTIONS – unchanged
# ============================================================

def reconstruct_abstract(inverted_index):
    if not inverted_index:
        return ""
    positions = {}
    for word, idxs in inverted_index.items():
        for i in idxs:
            positions[i] = word
    if not positions:
        return ""
    max_pos = max(positions.keys())
    return " ".join(positions.get(i, "") for i in range(max_pos + 1))

def extract_authors(work):
    names = []
    for a in work.get("authorships", []):
        author = a.get("author", {}) or {}
        name = author.get("display_name")
        if name:
            names.append(name)
    return "; ".join(names)

def extract_journal(work):
    primary = work.get("primary_location") or {}
    source = primary.get("source") or {}
    return source.get("display_name", "") or ""

def extract_pdf_url(work):
    for loc_key in ("best_oa_location", "primary_location"):
        loc = work.get(loc_key) or {}
        pdf = loc.get("pdf_url")
        if pdf:
            return pdf
    open_access = work.get("open_access") or {}
    return open_access.get("oa_url", "") or ""

def normalize_work(work):
    doi = work.get("doi", "") or ""
    if doi.startswith("https://doi.org/"):
        doi_url = doi
    elif doi:
        doi_url = f"https://doi.org/{doi}"
    else:
        doi_url = ""
    return {
        "id": work.get("id", ""),
        "title": work.get("display_name", "") or work.get("title", "") or "",
        "authors": extract_authors(work),
        "journal": extract_journal(work),
        "abstract": reconstruct_abstract(work.get("abstract_inverted_index")),
        "doi_url": doi_url,
        "pdf_url": extract_pdf_url(work),
        "year": work.get("publication_year"),
    }

OPENALEX_URL = "https://api.openalex.org/works"

# ============================================================
# 2. SEARCH FUNCTION – with optional expanded index
# ============================================================

def search_openalex_phrase_year(phrase, year, email=None, api_key=None, per_page=200,
                                session=None, sleep_between=0.1, work_types=None,
                                include_expanded=False, just_count=False):
    session = session or requests.Session()
    cursor = "*"
    count = 0

    filter_parts = [f"publication_year:{year}"]
    if work_types and "All types" not in work_types:
        type_filter = "|".join(work_types)
        filter_parts.append(f"type:{type_filter}")
    filter_string = ",".join(filter_parts)

    params = {
        "search": phrase,
        "filter": filter_string,
        "per-page": 1 if just_count else per_page,
    }
    if include_expanded:
        params["corpus"] = "all"

    if email:
        params["mailto"] = email
    if api_key:
        params["api_key"] = api_key

    temp_params = dict(params)
    temp_params["cursor"] = "*"
    debug_url = f"{OPENALEX_URL}?{urllib.parse.urlencode(temp_params)}"

    if just_count:
        try:
            resp = session.get(OPENALEX_URL, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            count = data.get("meta", {}).get("count", 0)
        except requests.exceptions.HTTPError as e:
            st.error(f"❌ OpenAlex API error for count of '{phrase}' (year {year}):")
            st.error(f"Status code: {resp.status_code}")
            st.error(f"Response body (first 500 chars):\n{resp.text[:500]}")
            raise
        return count, debug_url
    else:
        results = []
        first_page_meta_count = None
        while cursor:
            params["cursor"] = cursor
            try:
                resp = session.get(OPENALEX_URL, params=params, timeout=30)
                resp.raise_for_status()
            except requests.exceptions.HTTPError as e:
                st.error(f"❌ OpenAlex API error for phrase '{phrase}' (year {year}):")
                st.error(f"Status code: {resp.status_code}")
                st.error(f"Response body (first 500 chars):\n{resp.text[:500]}")
                raise

            data = resp.json()
            if first_page_meta_count is None:
                first_page_meta_count = data.get("meta", {}).get("count", 0)

            for work in data.get("results", []):
                results.append(normalize_work(work))

            cursor = (data.get("meta", {}) or {}).get("next_cursor")
            if not data.get("results"):
                break
            time.sleep(sleep_between)
        return results, first_page_meta_count, debug_url

# ============================================================
# 3. COLLECT FUNCTION
# ============================================================

def collect(phrases, years, email=None, api_key=None, sleep_between=0.1, work_types=None,
            include_expanded=False, fetch_fn=search_openalex_phrase_year):
    by_year = {y: {} for y in years}
    debug_info = {}

    for year in years:
        for phrase in phrases:
            rows, count, debug_url = fetch_fn(
                phrase, year, email=email, api_key=api_key,
                sleep_between=sleep_between, work_types=work_types,
                include_expanded=include_expanded,
                just_count=False
            )
            debug_info[(phrase, year)] = {
                "count": count,
                "url": debug_url
            }
            for row in rows:
                key = row["id"] or row["doi_url"] or f"{row['title'].strip().lower()}|{row['year']}"
                if key in by_year[year]:
                    existing = by_year[year][key]
                    if phrase not in existing["matched_phrases"]:
                        existing["matched_phrases"].append(phrase)
                else:
                    row = dict(row)
                    row["matched_phrases"] = [phrase]
                    by_year[year][key] = row
    return {year: list(rows.values()) for year, rows in by_year.items()}, debug_info

# ============================================================
# 4. EXPORT FUNCTIONS – with filters in settings
# ============================================================

def sanitize_for_excel(value):
    if isinstance(value, str):
        return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', value)
    return value

def write_excel_sheet(ws, rows, header, bold):
    for col_idx, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=sanitize_for_excel(row["title"]))
        ws.cell(row=r_idx, column=2, value=sanitize_for_excel(row["authors"]))
        ws.cell(row=r_idx, column=3, value=sanitize_for_excel(row["journal"]))
        ws.cell(row=r_idx, column=4, value=sanitize_for_excel(row["abstract"]))
        doi_cell = ws.cell(row=r_idx, column=5, value=row["doi_url"])
        if row["doi_url"]:
            doi_cell.hyperlink = row["doi_url"]
            doi_cell.font = Font(color="0563C1", underline="single")
        pdf_cell = ws.cell(row=r_idx, column=6, value=row["pdf_url"])
        if row["pdf_url"]:
            pdf_cell.hyperlink = row["pdf_url"]
            pdf_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=r_idx, column=7, value="; ".join(row.get("matched_phrases", [])))
    for i, w in enumerate([50, 30, 30, 70, 35, 45, 25], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row > 1:
        table = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=f"A1:{get_column_letter(max_col)}{max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

def export_to_excel_bytes(merged_rows, all_phrases, all_years, email=None, separate_tabs=False,
                          filters=None):
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header = ["Title", "Authors", "Journal", "Abstract", "DOI", "PDF URL", "Matched Phrase(s)"]
    bold = Font(bold=True)
    missing_pdfs_rows = []

    if separate_tabs:
        by_year = {}
        for year, row in merged_rows:
            by_year.setdefault(year, []).append(row)
        for year in sorted(by_year.keys()):
            ws = wb.create_sheet(title=str(year))
            rows = sorted(by_year[year], key=lambda r: r["title"].lower())
            write_excel_sheet(ws, rows, header, bold)
            for r in rows:
                if not r.get("pdf_url"):
                    missing_pdfs_rows.append(r)
    else:
        ws = wb.create_sheet(title="All Records")
        rows = sorted([r[1] for r in merged_rows], key=lambda r: (r["year"], r["title"].lower()))
        write_excel_sheet(ws, rows, header, bold)
        for r in rows:
            if not r.get("pdf_url"):
                missing_pdfs_rows.append(r)

    settings_ws = wb.create_sheet(title="Search Settings")
    settings_ws.cell(row=1, column=1, value="Setting").font = bold
    settings_ws.cell(row=1, column=2, value="Value").font = bold
    settings_ws.cell(row=2, column=1, value="Queries searched")
    settings_ws.cell(row=2, column=2, value="; ".join(all_phrases))
    settings_ws.cell(row=3, column=1, value="Publication years")
    settings_ws.cell(row=3, column=2, value=", ".join(str(y) for y in all_years))
    settings_ws.cell(row=4, column=1, value="OpenAlex email (polite pool)")
    settings_ws.cell(row=4, column=2, value=email or "Not provided")
    settings_ws.cell(row=5, column=1, value="API Key used")
    settings_ws.cell(row=5, column=2, value="Yes (server-side)" if st.secrets.get("OPENALEX_API_KEY") else "No")
    settings_ws.cell(row=6, column=1, value="Total included records")
    settings_ws.cell(row=6, column=2, value=len(merged_rows))
    row_num = 7
    if filters:
        for key, val in filters.items():
            if val.strip():
                settings_ws.cell(row=row_num, column=1, value=f"Filter: {key}")
                settings_ws.cell(row=row_num, column=2, value=val)
                row_num += 1
    settings_ws.column_dimensions["A"].width = 30
    settings_ws.column_dimensions["B"].width = 60

    missing_ws = wb.create_sheet(title="Missing PDFs")
    missing_headers = ["Title", "Authors", "Year", "DOI"]
    for col_idx, h in enumerate(missing_headers, start=1):
        missing_ws.cell(row=1, column=col_idx, value=h).font = bold
    if missing_pdfs_rows:
        for r_idx, r in enumerate(sorted(missing_pdfs_rows, key=lambda r: (r["year"], r["title"].lower())), start=2):
            missing_ws.cell(row=r_idx, column=1, value=sanitize_for_excel(r["title"]))
            missing_ws.cell(row=r_idx, column=2, value=sanitize_for_excel(r["authors"]))
            missing_ws.cell(row=r_idx, column=3, value=r["year"])
            doi_cell = missing_ws.cell(row=r_idx, column=4, value=r["doi_url"])
            if r["doi_url"]:
                doi_cell.hyperlink = r["doi_url"]
                doi_cell.font = Font(color="0563C1", underline="single")
    else:
        missing_ws.cell(row=2, column=1, value="No missing PDFs in the download list.")
    for i, w in enumerate([50, 30, 10, 35], start=1):
        missing_ws.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_to_csv_bytes(merged_rows):
    data = []
    for year, r in merged_rows:
        data.append({
            "Year": year,
            "Title": r["title"],
            "Authors": r["authors"],
            "Journal": r["journal"],
            "Abstract": r["abstract"],
            "DOI": r["doi_url"],
            "PDF URL": r["pdf_url"],
            "Matched Phrases": "; ".join(r.get("matched_phrases", []))
        })
    df = pd.DataFrame(data)
    df = df.sort_values(["Year", "Title"]).reset_index(drop=True)
    out = StringIO()
    df.to_csv(out, index=False)
    return out.getvalue().encode('utf-8')

# ============================================================
# 5. PRE‑FLIGHT COUNT
# ============================================================

def get_total_count(phrases, years, email, api_key, work_types, include_expanded, fetch_fn):
    total = 0
    for year in years:
        for phrase in phrases:
            count, _ = fetch_fn(phrase, year, email=email, api_key=api_key,
                                work_types=work_types, include_expanded=include_expanded,
                                just_count=True)
            total += count
    return total

# ============================================================
# 6. STREAMLIT UI – compact layout with filters in right column
# ============================================================

st.set_page_config(page_title="LitFind", layout="wide")

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 3.5rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

api_key = st.secrets.get("OPENALEX_API_KEY")
if not api_key:
    st.error("🚨 **API Key Missing!** Please add OPENALEX_API_KEY to your secrets.")
    st.stop()

st.markdown(
    """
    <div style="display: flex; align-items: center; gap: 0.7rem; margin-bottom: 0.2rem;">
        <svg width="38" height="38" viewBox="0 0 40 40" style="flex-shrink: 0;">
            <mask id="searchMask">
                <rect width="40" height="40" fill="white"/>
                <circle cx="17" cy="17" r="8" fill="none" stroke="black" stroke-width="3.5"/>
                <line x1="23" y1="23" x2="31" y2="31" stroke="black" stroke-width="3.5" stroke-linecap="round"/>
            </mask>
            <rect width="40" height="40" rx="9" fill="#2563eb" mask="url(#searchMask)"/>
        </svg>
        <span style="font-size: 1.9rem; font-weight: 700; letter-spacing: -0.01em; line-height: 1.3; color: #1a1a2e;">LitFind</span>
    </div>
    """,
    unsafe_allow_html=True,
)
st.markdown("Search for scholarly works using **OpenAlex**. Enter your search phrases and years to find relevant publications – abstracts, authors, journals and PDF links are all included.")

with st.expander("ℹ️ About this search tool", expanded=False):
    st.markdown("""
    This tool searches **OpenAlex** – a free, open index of the world's research ecosystem.

    **What does OpenAlex cover?**
    - Over **320 million scholarly works** in the default **Core** corpus: journal articles, conference papers, books, book chapters, datasets, dissertations, preprints, and more
    - An **optional Expanded index** (XPAC) with an additional **190 million** records from DataCite, institutional repositories, and preprints – often with sparser metadata (lower quality). You can enable it with the toggle below.

    **Why OpenAlex?**
    - It is **free and open** – no paywalls, no API keys required (though providing your email gives you faster "polite pool" access)
    - It is **more comprehensive** than Scopus or Web of Science, with over 464 million works indexed
    - It includes **datasets, software, and other research objects** beyond just traditional publications

    **Search tips:**
    - Use **uppercase** `AND`, `OR`, `NOT` for boolean logic
    - Use **double quotes** for exact phrase matches (e.g., `"climate change"`)
    - **Parentheses** group terms (e.g., `(neural OR deep)`)
    - Field‑specific syntax: `title:"blended learning"`, `author:"Smith"`, `journal:"Nature"`
    - Non‑ASCII characters (e.g., 守破離) are fully supported
    - You can filter results by document type using the dropdown below
    - Enable the **Expanded index** to include lower‑quality records from repositories – use with caution as it may increase noise

    **Optional filters**: Use the builder below to add constraints (Author, Title, Journal, Abstract, Institution). Each filter is combined with `AND`. They are applied to all queries.
    - The builder **automatically adds quotes** around your value to ensure exact‑phrase matching. You can still type your own quotes if you prefer.
    """)

# ---- Session state ----
if "email" not in st.session_state:
    st.session_state.email = ""
if "search_sessions" not in st.session_state:
    st.session_state.search_sessions = []
if "do_full_fetch" not in st.session_state:
    st.session_state.do_full_fetch = False
if "force_refresh" not in st.session_state:
    st.session_state.force_refresh = False
if "download_rows" not in st.session_state:
    st.session_state.download_rows = []
if "filters_list" not in st.session_state:
    st.session_state.filters_list = []  # list of {"field": "author", "value": "Smith"}

# ---- LAYOUT: main form on left, extra controls on right ----
col_left, col_right = st.columns([1, 1])

# ---- LEFT COLUMN: Search form ----
with col_left:
    with st.form("search_form"):
        phrases_input = st.text_area(
            "🔍 Search Queries",
            value="",
            placeholder="Enter your search queries here, one per line. Example: title:\"blended learning\" AND author:\"Smith\"",
            help="Each line is a separate query. Use boolean operators and field prefixes."
        )
        start_year = st.number_input("Start Year", min_value=1900, max_value=2030, value=2020, step=1)
        end_year = st.number_input("End Year", min_value=1900, max_value=2030, value=2026, step=1)
        force_refresh = st.checkbox("🔄 Force Refresh (Ignore Cache)", value=False)

        # The Run Search button must be inside the form
        submitted = st.form_submit_button("🚀 Run Search", use_container_width=True)

# ---- RIGHT COLUMN: fields that live next to the form, plus filter builder and Clear ----
with col_right:
    # 1. Email (outside the form so it's not cleared on submit)
    email = st.text_input(
        "📧 Recommended: Your Email (for OpenAlex polite pool)",
        value=st.session_state.email,
        help="Optional, but using a real email gives you 10x more daily searches."
    )
    st.caption("**Providing an email is optional** – without it, you'll have a lower daily quota.")

    # 2. Work Types
    work_type_options = ["All types", "article", "book", "book-chapter", "dataset",
                         "dissertation", "preprint", "conference-paper", "conference-abstract",
                         "book-review", "report", "editorial", "letter", "erratum"]
    work_types = st.multiselect(
        "📚 Work Types (optional – select multiple)",
        options=work_type_options,
        default=["All types"],
        help="Filter by document type. 'All types' means no filter."
    )
    if "All types" in work_types:
        work_types = ["All types"]

    # 3. Expanded index toggle
    include_expanded = st.checkbox(
        "🌐 Include expanded index (XPAC)",
        value=False,
        help="Adds ~190M lower‑quality records from DataCite and repositories. Increases recall but may add noise."
    )

    # 4. Optional Filters builder (outside the form, below the toggles)
    st.markdown("---")
    st.markdown("**🔍 Optional Filters**")
    st.caption("Add constraints combined with AND. Quotes are added automatically for exact match.")

    # Layout for adding a filter
    col_field, col_value, col_add = st.columns([2, 3, 1])
    with col_field:
        field_options = ["Author", "Title", "Journal", "Abstract", "Institution"]
        selected_field = st.selectbox("Field", options=field_options, key="filter_field", label_visibility="collapsed")
    with col_value:
        filter_value = st.text_input("Value", key="filter_value", placeholder="e.g., Smith or \"John Smith\"", label_visibility="collapsed")
    with col_add:
        if st.button("➕ Add", use_container_width=True):
            if filter_value.strip():
                st.session_state.filters_list.append({"field": selected_field.lower(), "value": filter_value.strip()})
                st.rerun()

    # Display current filters with remove buttons
    if st.session_state.filters_list:
        for i, f in enumerate(st.session_state.filters_list):
            col_f, col_v, col_del = st.columns([2, 3, 1])
            with col_f:
                st.caption(f"**{f['field'].capitalize()}**")
            with col_v:
                st.caption(f"`{f['value']}`")
            with col_del:
                if st.button("✕", key=f"del_filter_{i}"):
                    del st.session_state.filters_list[i]
                    st.rerun()
    else:
        st.caption("*No active filters*")

    # 5. Clear All button (below filters)
    st.markdown("---")
    if st.button("🗑️ Clear All Searches", use_container_width=True):
        st.session_state.search_sessions = []
        st.session_state.download_rows = []
        st.session_state.filters_list = []
        for key in list(st.session_state.keys()):
            if key.startswith("df_"):
                del st.session_state[key]
        st.rerun()

# ---- Process the form submission ----
if submitted:
    base_queries = [p.strip() for p in phrases_input.splitlines() if p.strip()]
    if not base_queries:
        st.error("Please enter at least one search query.")
        st.stop()

    # Build filter string from the current filters_list, auto‑quoting values
    filter_parts = []
    for f in st.session_state.filters_list:
        field = f["field"]
        value = f["value"]
        # Add quotes if not already present (for exact phrase matching)
        if not (value.startswith('"') and value.endswith('"')):
            value = f'"{value}"'
        filter_parts.append(f"{field}:{value}")
    filters_str = " AND ".join(filter_parts) if filter_parts else ""

    # Build full queries: each base query + AND filters if any
    full_queries = []
    for q in base_queries:
        if filters_str:
            full_queries.append(f"{q} AND {filters_str}")
        else:
            full_queries.append(q)

    years = list(range(int(start_year), int(end_year) + 1))
    st.session_state.email = email
    st.session_state.base_queries = base_queries
    st.session_state.full_queries = full_queries
    st.session_state.years = years
    st.session_state.work_types = work_types
    st.session_state.force_refresh = force_refresh
    st.session_state.include_expanded = include_expanded
    # Store filters as dict for display in settings
    filters_dict = {f["field"].capitalize(): f["value"] for f in st.session_state.filters_list}
    st.session_state.filters = filters_dict

    with st.status("🔎 Checking search scope...", expanded=True) as status:
        try:
            total_count = get_total_count(
                full_queries, years, email, api_key, work_types, include_expanded,
                search_openalex_phrase_year
            )
            st.session_state.preflight_count = total_count
            status.update(label=f"🔎 Found approximately {total_count} matching works.", state="running")
        except Exception as e:
            st.error(f"Pre‑flight count failed: {e}")
            st.stop()

    THRESHOLD = 2000
    if total_count > THRESHOLD:
        st.warning(f"⚠️ **Search too broad!** ~{total_count} works. Please narrow your search.")
        force_check = st.checkbox("⚠️ **Force full fetch anyway** (I understand the risks)")
        if force_check:
            if st.button("📥 Fetch all records"):
                st.session_state.do_full_fetch = True
                st.rerun()
        else:
            st.stop()
    else:
        st.session_state.do_full_fetch = True
        st.rerun()

# ---- Full fetch (cached) ----
if st.session_state.do_full_fetch:
    full_queries = st.session_state.full_queries
    base_queries = st.session_state.base_queries
    years = st.session_state.years
    work_types = st.session_state.work_types
    email = st.session_state.email
    force_refresh = st.session_state.force_refresh
    include_expanded = st.session_state.include_expanded
    filters = st.session_state.filters

    @st.cache_data(show_spinner=False)
    def run_collection_cached(queries_tuple, years_tuple, email, api_key, work_types_tuple,
                              include_expanded, refresh_seed):
        return collect(list(queries_tuple), list(years_tuple), email=email, api_key=api_key,
                       work_types=list(work_types_tuple), include_expanded=include_expanded)

    with st.status("⏳ Fetching full records...", expanded=True) as status:
        refresh_seed = random.randint(0, 999999) if force_refresh else 0
        results_by_year, _ = run_collection_cached(
            tuple(full_queries), tuple(years), email, api_key, tuple(work_types),
            include_expanded, refresh_seed
        )
        seen_keys = set()
        flat_rows = []
        for year, rows in results_by_year.items():
            for r in rows:
                key = r.get("id") or r.get("doi_url") or f"{r['title'].strip().lower()}|{r['year']}"
                if key not in seen_keys:
                    seen_keys.add(key)
                    flat_rows.append((year, r))
        total = len(flat_rows)
        status.update(label=f"✅ Fetched {total} records.", state="complete")

    new_search = {
        "id": datetime.now().strftime("%Y%m%d_%H%M%S"),
        "base_queries": base_queries,
        "full_queries": full_queries,
        "years": years,
        "work_types": work_types,
        "include_expanded": include_expanded,
        "filters": filters,
        "flat_rows": flat_rows,
        "total": total,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "included_indices": set()
    }
    st.session_state.search_sessions.insert(0, new_search)
    st.session_state.do_full_fetch = False
    st.rerun()

# ---- Display search sessions ----
if st.session_state.search_sessions:
    st.subheader("📋 All Searches (newest first)")
    st.caption("For each search, tick the rows you want to include. Then click the **Apply** button at the bottom to build the download list.")

    for idx, sess in enumerate(st.session_state.search_sessions):
        flat_rows = sess["flat_rows"]
        total = sess["total"]
        if total == 0:
            continue
        included_indices = sess.get("included_indices", set())

        display_queries = sess.get("base_queries", sess.get("full_queries", []))
        filters_display = sess.get("filters", {})
        filter_str = ""
        if filters_display:
            active = [f"{k}:{v}" for k, v in filters_display.items() if v.strip()]
            if active:
                filter_str = " [filters: " + ", ".join(active) + "]"

        all_rows = []
        for i, (year, r) in enumerate(flat_rows):
            all_rows.append({
                "Include": i in included_indices,
                "Year": year,
                "Title": r["title"],
                "Authors": r["authors"],
                "Journal": r["journal"],
                "Abstract": r["abstract"],
                "Has PDF": "✅" if r["pdf_url"] else "❌",
                "Phrases": "; ".join(r.get("matched_phrases", []))
            })
        df = pd.DataFrame(all_rows)

        editor_key = f"include_editor_{sess['id']}"
        if f"df_{sess['id']}" not in st.session_state:
            st.session_state[f"df_{sess['id']}"] = df

        with st.expander(f"Search {idx+1}: {display_queries[0][:60]}…{filter_str} ({total} records, {len(included_indices)} included)", expanded=(idx < 2)):
            colA, colB = st.columns([1, 1])
            if colA.button(f"✅ Include All", key=f"include_all_{sess['id']}"):
                df_local = st.session_state[f"df_{sess['id']}"].copy()
                df_local["Include"] = True
                st.session_state[f"df_{sess['id']}"] = df_local
                st.rerun()
            if colB.button(f"❌ Include None", key=f"include_none_{sess['id']}"):
                df_local = st.session_state[f"df_{sess['id']}"].copy()
                df_local["Include"] = False
                st.session_state[f"df_{sess['id']}"] = df_local
                st.rerun()

            edited_df = st.data_editor(
                st.session_state[f"df_{sess['id']}"],
                use_container_width=True,
                height=400,
                column_config={
                    "Include": st.column_config.CheckboxColumn(
                        "Include",
                        width=80,
                        help="Check to include this row in the download list"
                    ),
                    "Year": st.column_config.NumberColumn("Year", width="small"),
                    "Title": st.column_config.TextColumn("Title", width="large"),
                    "Authors": st.column_config.TextColumn("Authors", width="medium"),
                    "Journal": st.column_config.TextColumn("Journal", width="medium"),
                    "Abstract": st.column_config.TextColumn(
                        "Abstract (double click to expand)",
                        width="large",
                        disabled=False
                    ),
                    "Has PDF": st.column_config.TextColumn("PDF", width="small"),
                    "Phrases": st.column_config.TextColumn("Matched Phrases", width="medium"),
                },
                hide_index=True,
                key=editor_key
            )
            new_included = set()
            for row_idx, row in edited_df.iterrows():
                if row["Include"]:
                    new_included.add(row_idx)
            sess["included_indices"] = new_included
            st.session_state[f"df_{sess['id']}"] = edited_df
            st.caption(f"📌 {len(new_included)} rows included from this search.")

    if st.button("🔄 Apply Selected Records", use_container_width=True):
        all_included = []
        for sess in st.session_state.search_sessions:
            included = sess.get("included_indices", set())
            for row_idx in included:
                if row_idx < len(sess["flat_rows"]):
                    all_included.append(sess["flat_rows"][row_idx])
        seen = set()
        download_rows = []
        for year, r in all_included:
            key = r.get("id") or r.get("doi_url") or f"{r['title'].strip().lower()}|{r['year']}"
            if key not in seen:
                seen.add(key)
                download_rows.append((year, r))
        st.session_state.download_rows = download_rows
        st.success(f"✅ Download list updated: {len(download_rows)} unique records.")
        st.rerun()

    if st.session_state.download_rows:
        download_rows = st.session_state.download_rows
        st.subheader(f"📦 Download List – {len(download_rows)} unique included records")
        st.caption("This is the union of all rows you marked 'Include' across all searches. Download this set below.")

        preview_data = []
        for i, (year, r) in enumerate(download_rows, start=1):
            preview_data.append({
                "No.": i,
                "Year": year,
                "Title": r["title"]
            })
        preview_df = pd.DataFrame(preview_data)
        st.dataframe(preview_df, use_container_width=True, height=200)

        # ---- Download section ----
        col_name, col_format, col_layout, col_btn = st.columns([2, 1, 1, 1])
        with col_name:
            if st.session_state.search_sessions:
                first_base = st.session_state.search_sessions[0].get("base_queries", [""])[0]
                summary = "_".join([first_base])[:50].replace(" ", "_").replace('"', '').replace("'", "")
                summary = re.sub(r'[^a-zA-Z0-9_]', '', summary)
                base_name = f"{datetime.now().strftime('%Y-%m-%d')}_{summary}"
            else:
                base_name = f"{datetime.now().strftime('%Y-%m-%d')}_download"
            filename = st.text_input(
                "📁 Filename (without extension)",
                value=base_name,
                help="Customise the base file name before downloading.",
                key="download_filename"
            )
        with col_format:
            file_format = st.radio(
                "Format",
                options=["Excel", "CSV"],
                index=0,
                key="file_format"
            )
        with col_layout:
            if file_format == "Excel":
                export_mode = st.radio(
                    "Layout",
                    options=["Single sheet", "Separate tabs by year"],
                    index=0,
                    key="export_mode"
                )
            else:
                export_mode = "Single sheet"
                st.write("")  # placeholder
        with col_btn:
            st.write("")  # vertical spacer
            st.write("")
            if st.button("⬇️ Download File", use_container_width=True, key="download_btn"):
                all_base_queries = []
                all_years = []
                all_filters = {}
                for sess in st.session_state.search_sessions:
                    all_base_queries.extend(sess.get("base_queries", sess.get("full_queries", [])))
                    all_years.extend(sess["years"])
                    if sess.get("filters"):
                        all_filters = sess["filters"]
                all_base_queries = list(dict.fromkeys(all_base_queries))
                all_years = sorted(set(all_years))

                if file_format == "Excel":
                    ext = ".xlsx"
                    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    output = export_to_excel_bytes(
                        download_rows,
                        all_base_queries,
                        all_years,
                        st.session_state.email,
                        separate_tabs=(export_mode == "Separate tabs by year"),
                        filters=all_filters
                    )
                else:
                    ext = ".csv"
                    mime = "text/csv"
                    output = export_to_csv_bytes(download_rows)

                st.download_button(
                    label="📥 Click to save",
                    data=output,
                    file_name=f"{filename.strip() if filename.strip() else base_name}{ext}",
                    mime=mime,
                    use_container_width=True,
                    key="final_download"
                )
    else:
        st.info("No records in download list yet. Mark rows as 'Include' in any search and click Apply.")
else:
    st.info("No searches yet. Fill in the form above and click 'Run Search'.")
